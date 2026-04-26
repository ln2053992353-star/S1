import os
import sys
import json
import csv
import time
from datetime import datetime
from django.core.management.base import BaseCommand
from django.db import reset_queries
from django.conf import settings

# 尝试导入openpyxl用于Excel导出
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 设置 Hugging Face 国内镜像（用户要求）
os.environ["HF_ENDPOINT"] = "https://hf-mirror.com"

from search_engine.models import Product, ProductEmbedding, Tag, PubChemTag, ProductSource


class Command(BaseCommand):
    help = '全量导出重构后的 Product 及 ProductEmbedding 数据（零性能隐患版本）'

    def add_arguments(self, parser):
        """定义命令行参数"""
        parser.add_argument(
            '--format',
            type=str,
            default='jsonl',
            choices=['jsonl', 'csv', 'excel'],
            help='输出格式：jsonl (JSON Lines)、csv 或 excel，默认 jsonl'
        )
        parser.add_argument(
            '--output',
            type=str,
            default='product_export',
            help='输出文件路径前缀（不含扩展名），默认 product_export'
        )
        parser.add_argument(
            '--chunk-size',
            type=int,
            default=1000,
            help='分块大小（基于主键的游标分页），默认 1000'
        )
        parser.add_argument(
            '--include-vector',
            action='store_true',
            help='是否包含向量数据（vector 字段）'
        )
        parser.add_argument(
            '--min-id',
            type=int,
            default=1,
            help='起始产品ID（用于断点续传），默认 1'
        )

    def handle(self, *args, **options):
        """命令主处理逻辑"""
        self.stdout.write("开始全量导出重构数据（零性能隐患版本）")
        self.stdout.write("=" * 60)

        # 解析参数
        output_format = options['format']
        output_prefix = options['output']
        chunk_size = options['chunk_size']
        include_vector = options['include_vector']
        min_id = options['min_id']

        # 存储为实例变量，供其他方法使用
        self.include_vector = include_vector

        # 设置输出文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if output_format == 'jsonl':
            output_file = f"{output_prefix}_{timestamp}.jsonl"
        elif output_format == 'csv':
            output_file = f"{output_prefix}_{timestamp}.csv"
        else:  # excel
            if not OPENPYXL_AVAILABLE:
                self.stderr.write("❌ 错误: openpyxl库未安装，无法生成Excel文件")
                self.stderr.write("请安装: pip install openpyxl")
                sys.exit(1)
            output_file = f"{output_prefix}_{timestamp}.xlsx"

        self.stdout.write(f"输出文件: {output_file}")
        self.stdout.write(f"输出格式: {output_format}")
        self.stdout.write(f"分块大小: {chunk_size}")
        self.stdout.write(f"包含向量: {'是' if include_vector else '否'}")
        self.stdout.write(f"起始ID: {min_id}")
        self.stdout.write("=" * 60)

        # 记录开始时间
        start_time = time.time()
        total_processed = 0
        total_failed = 0
        last_processed_id = min_id - 1

        try:
            # 根据输出格式初始化写入器
            writer = None
            wb = None
            ws = None
            fieldnames = None
            excel_row_counter = 1  # 从1开始，第一行是表头
            f = None
            csv_writer = None

            if output_format == 'excel':
                # 初始化Excel写入器
                wb, ws, fieldnames = self._initialize_excel_writer()
                excel_row_counter = 2  # 表头在第1行，数据从第2行开始
            elif output_format in ['jsonl', 'csv']:
                # 对于JSONL和CSV，使用文件流式写入
                f = open(output_file, 'w', encoding='utf-8')
                if output_format == 'csv':
                    # 初始化 CSV writer 并写入表头
                    csv_writer = self._initialize_csv_writer(f)

            # 主循环：基于主键的游标分页（Keyset Pagination）
            while True:
                # 🔴 性能红线：使用 select_related + prefetch_related 避免 N+1 查询
                # 🔴 工程防线：使用 filter(id__gt=last_id) 避免 OFFSET 深度分页
                # 🔴 数据库兼容：使用 defer 排除不存在的列，防止查询失败
                products = Product.objects.select_related(
                    'embedding',
                    'pubchem_data'
                ).prefetch_related(
                    'tags',
                    'pubchem_tags',
                    'sources'
                ).filter(
                    product_id__gt=last_processed_id,
                    embedding__isnull=False  # 只导出有向量的产品
                ).defer(
                    'embedding__function',
                    'embedding__pubchem_description',
                    'embedding__tags_text',
                    'embedding__grna',
                    'embedding__iupac_name',
                    'embedding__source_database'
                ).order_by('product_id')[:chunk_size]

                # 转换为列表以评估查询
                chunk_list = list(products)
                if not chunk_list:
                    break  # 没有更多数据

                # 处理当前分块
                chunk_start_time = time.time()
                chunk_processed = 0
                chunk_failed = 0

                for product in chunk_list:
                    try:
                        # 构建导出记录
                        record = self._build_product_record(product)

                        # 根据输出格式写入数据
                        if output_format == 'excel':
                            # Excel格式：写入到工作表中
                            for col_idx, field in enumerate(fieldnames, 1):
                                value = record.get(field, '')
                                ws.cell(row=excel_row_counter, column=col_idx, value=value)
                            excel_row_counter += 1
                        elif output_format == 'jsonl':
                            # 🔴 工程防线：JSONL 流式写入，避免内存中构建巨大 JSON 数组
                            f.write(json.dumps(record, ensure_ascii=False) + '\n')
                        elif output_format == 'csv':
                            csv_writer.writerow(record)

                        chunk_processed += 1
                        last_processed_id = product.product_id

                    except Exception as e:
                        chunk_failed += 1
                        self.stderr.write(
                            f"❌ 导出失败 Product ID {product.product_id}: {str(e)}"
                        )

                # 更新统计
                total_processed += chunk_processed
                total_failed += chunk_failed

                # 显示进度
                chunk_time = time.time() - chunk_start_time
                self._show_progress(
                    last_processed_id, total_processed, total_failed,
                    chunk_processed, chunk_time
                )

                # 🔴 工程防线：清理 Django SQL 查询日志，防止 DEBUG 模式内存泄漏
                reset_queries()

            # 完成写入，关闭文件
            if output_format == 'excel':
                # 保存Excel文件
                wb.save(output_file)
                self.stdout.write(f"Excel文件已保存: {output_file}")
            elif output_format in ['jsonl', 'csv']:
                f.close()
                self.stdout.write(f"文件已保存: {output_file}")

            # 计算总耗时
            total_time = time.time() - start_time

            # 输出最终统计
            self.stdout.write("\n" + "=" * 60)
            self.stdout.write("导出完成！")
            self.stdout.write(f"总处理记录: {total_processed}")
            self.stdout.write(f"失败记录: {total_failed}")
            self.stdout.write(f"总耗时: {total_time:.2f} 秒")
            if total_processed > 0:
                self.stdout.write(f"平均速度: {total_processed/total_time:.2f} 条/秒")
            self.stdout.write(f"输出文件: {output_file}")
            self.stdout.write("=" * 60)

            # 如果使用了断点续传，提示下次可以使用的起始ID
            if last_processed_id > 0:
                self.stdout.write(
                    f"如需继续导出，可使用参数: --min-id {last_processed_id + 1}"
                )

        except KeyboardInterrupt:
            self.stdout.write("\n用户中断导出")
            self.stdout.write(f"最后处理的 Product ID: {last_processed_id}")
            # 尝试保存Excel文件（如果已初始化）
            if output_format == 'excel' and wb is not None:
                try:
                    wb.save(output_file)
                    self.stdout.write(f"已保存部分Excel文件: {output_file}")
                except Exception as save_err:
                    self.stderr.write(f"保存Excel文件失败: {save_err}")
            # 关闭文件（如果已打开）
            elif output_format in ['jsonl', 'csv'] and f is not None:
                try:
                    f.close()
                except Exception as close_err:
                    self.stderr.write(f"关闭文件失败: {close_err}")

            self.stdout.write(f"如需继续导出，可使用参数: --min-id {last_processed_id + 1}")
            sys.exit(1)
        except Exception as e:
            self.stderr.write(f"导出过程发生致命错误: {str(e)}")
            # 确保关闭文件
            if output_format == 'excel' and wb is not None:
                try:
                    wb.save(output_file)
                except Exception:
                    pass  # 忽略保存错误
            elif output_format in ['jsonl', 'csv'] and f is not None:
                try:
                    f.close()
                except Exception:
                    pass  # 忽略关闭错误
            sys.exit(1)
        finally:
            # 确保文件被关闭（用于正常流程结束）
            if output_format in ['jsonl', 'csv'] and f is not None:
                try:
                    f.close()
                except Exception:
                    pass  # 忽略关闭错误

    def _initialize_csv_writer(self, file_obj):
        """初始化 CSV writer 并写入表头"""
        # 定义 CSV 字段顺序
        fieldnames = self._get_fieldnames()

        writer = csv.DictWriter(file_obj, fieldnames=fieldnames)
        writer.writeheader()
        return writer

    def _initialize_excel_writer(self):
        """初始化 Excel writer 并写入表头"""
        # 创建Workbook和活跃工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"

        # 获取字段名
        fieldnames = self._get_fieldnames()

        # 写入表头（第一行）
        for col_idx, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col_idx, value=field)

        return wb, ws, fieldnames

    def _get_fieldnames(self):
        """获取字段名列表（CSV和Excel共享）"""
        fieldnames = [
            'product_id', 'product_name', 'grna_map', 'description',
            'source_filename', 'source_doi', 'tags', 'pubchem_tags',
            'sources', 'embedding_text', 'function', 'pubchem_description',
            'tags_text', 'grna', 'iupac_name', 'source_database',
            'model_name', 'dim', 'created_at', 'updated_at'
        ]

        if self.include_vector:
            fieldnames.append('vector')

        return fieldnames

    def _build_product_record(self, product):
        """构建单个产品的导出记录"""
        # 🔴 性能红线：直接从 prefetch 缓存中读取标签，禁止二次查库
        # 🔴 注意：使用 .all() 访问 prefetch_related 缓存
        tags = [tag.tag_name for tag in product.tags.all()]
        pubchem_tags = [tag.tag_name for tag in product.pubchem_tags.all()]
        sources = [source.doi for source in product.sources.all()]

        # 获取关联的 ProductEmbedding
        embedding = product.embedding

        # 构建基础记录（使用安全属性访问，兼容缺失的数据库字段）
        record = {
            'product_id': product.product_id,
            'product_name': product.product_name or '',
            'grna_map': product.grna_map or '',
            'description': product.description or '',
            'source_filename': product.source_filename or '',
            'source_doi': product.source_doi or '',
            'tags': '; '.join(tags) if tags else '',
            'pubchem_tags': '; '.join(pubchem_tags) if pubchem_tags else '',
            'sources': '; '.join(sources) if sources else '',
            'embedding_text': getattr(embedding, 'embedding_text', '') if embedding else '',
            'function': getattr(embedding, 'function', '') if embedding else '',
            'pubchem_description': getattr(embedding, 'pubchem_description', '') if embedding else '',
            'tags_text': getattr(embedding, 'tags_text', '') if embedding else '',
            'grna': getattr(embedding, 'grna', '') if embedding else '',
            'iupac_name': getattr(embedding, 'iupac_name', '') if embedding else '',
            'source_database': getattr(embedding, 'source_database', '') if embedding else '',
            'model_name': getattr(embedding, 'model_name', '') if embedding else '',
            'dim': getattr(embedding, 'dim', '') if embedding else '',
            'created_at': embedding.created_at.isoformat() if embedding and embedding.created_at else '',
            'updated_at': embedding.updated_at.isoformat() if embedding and embedding.updated_at else '',
        }

        # 可选：包含向量数据
        if self.include_vector and embedding and embedding.vector:
            try:
                # 解析 JSON 格式的向量数据
                vector_data = json.loads(embedding.vector)
                record['vector'] = json.dumps(vector_data, ensure_ascii=False)
            except Exception:
                record['vector'] = ''

        return record

    def _show_progress(self, last_id, total_processed, total_failed,
                      chunk_processed, chunk_time):
        """显示处理进度"""
        # 计算处理速度
        speed = chunk_processed / chunk_time if chunk_time > 0 else 0

        # 构建进度信息
        progress_info = (
            f"当前ID: {last_id:6d} | "
            f"累计: {total_processed:6d} 条 | "
            f"失败: {total_failed:3d} 条 | "
            f"速度: {speed:6.1f} 条/秒"
        )

        # 使用标准输出写入，确保实时显示
        self.stdout.write(progress_info)
        self.stdout.flush()


if __name__ == '__main__':
    # 允许直接运行脚本（开发调试用）
    import django
    django.setup()

    command = Command()
    command.run_from_argv(['manage.py', 'export_new_db'])